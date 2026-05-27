import io
from datetime import datetime, date
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from ..database import get_db
from ..models import Ticket, TicketService, UserModel
from .auth import require_admin

router = APIRouter(prefix="/api/admin/reports", tags=["admin-reports"])


@router.get("/services")
def list_services(
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: UserModel = Depends(require_admin),
):
    q = db.query(TicketService).join(Ticket)
    if date_from:
        q = q.filter(Ticket.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        q = q.filter(Ticket.created_at <= datetime.strptime(date_to, "%Y-%m-%d") + __import__('datetime').timedelta(days=1))
    services = q.order_by(TicketService.created_at.desc()).all()
    total_amount = sum(s.total or 0 for s in services)
    return {
        "items": [{
            "id": s.id, "service_name": s.service_name,
            "quantity": s.quantity, "price": s.price, "total": s.total,
            "ticket_id": s.ticket_id,
            "ticket_number": s.ticket.number if s.ticket else "",
            "created_at": s.created_at.isoformat(),
        } for s in services],
        "total_count": len(services),
        "total_amount": total_amount,
    }


@router.post("/services/export")
def export_services_report(
    format: str = Query("xlsx"),
    date_from: str | None = Query(None),
    date_to: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: UserModel = Depends(require_admin),
):
    q = db.query(TicketService).join(Ticket)
    if date_from:
        q = q.filter(Ticket.created_at >= datetime.strptime(date_from, "%Y-%m-%d"))
    if date_to:
        q = q.filter(Ticket.created_at <= datetime.strptime(date_to, "%Y-%m-%d") + __import__('datetime').timedelta(days=1))
    services = q.order_by(TicketService.created_at.desc()).all()

    total_amount = sum(s.total or 0 for s in services)

    if format == "xlsx":
        return _export_services_excel(services, total_amount)
    elif format == "pdf":
        return _export_services_pdf(services, total_amount)
    raise HTTPException(400, "Неверный формат. Используйте xlsx или pdf")


def _export_services_excel(services, total_amount) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Услуги"

    header_font = Font(name="DejaVu Sans", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Border(*(Side(style="thin"),) * 4)
    cell_font = Font(name="DejaVu Sans", size=10)

    headers = ["№", "Услуга", "Кол-во", "Цена", "Сумма", "Заявка", "Дата"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center"); c.border = thin

    for idx, s in enumerate(services, 1):
        ws.append([idx, s.service_name, s.quantity, s.price, s.total or 0,
                   s.ticket.number if s.ticket else "",
                   s.created_at.strftime("%d.%m.%Y %H:%M") if s.created_at else ""])
        for col in range(1, 8):
            ws.cell(row=idx + 1, column=col).font = cell_font
            ws.cell(row=idx + 1, column=col).border = thin

    ws.append([])
    ws.append(["", "", "", "ИТОГО:", total_amount, "", ""])
    ws.cell(row=len(services) + 3, column=5).font = Font(name="DejaVu Sans", bold=True, size=11)

    for col, w in enumerate([5, 40, 8, 10, 12, 15, 18], 1):
        ws.column_dimensions[chr(64 + col)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"services_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_services_pdf(services, total_amount) -> Response:
    items_html = ""
    for idx, s in enumerate(services, 1):
        items_html += f"""<tr>
            <td>{idx}</td><td>{s.service_name}</td><td>{s.quantity}</td>
            <td>{s.price:.2f}</td><td>{(s.total or 0):.2f}</td>
            <td>{s.ticket.number if s.ticket else ''}</td>
            <td>{s.created_at.strftime('%d.%m.%Y') if s.created_at else ''}</td>
        </tr>"""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  @page {{ margin: 15mm; }}
  body {{ font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 12px; }}
  h2 {{ text-align: center; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #000; padding: 5px 7px; text-align: left; font-size: 10px; }}
  th {{ background: #4472C4; color: #fff; }}
  .total {{ text-align: right; font-size: 14px; font-weight: bold; margin-top: 10px; }}
  .date {{ text-align: right; font-size: 10px; color: #666; margin-bottom: 15px; }}
</style></head><body>
  <h2>Отчёт по услугам</h2>
  <div class="date">Сформирован: {datetime.utcnow().strftime('%d.%m.%Y %H:%M')}</div>
  <table>
    <tr><th>№</th><th>Услуга</th><th>Кол</th><th>Цена</th><th>Сумма</th><th>Заявка</th><th>Дата</th></tr>
    {items_html}
  </table>
  <div class="total">Итого: {total_amount:.2f} руб.</div>
  <p style="margin-top:30px; font-size:10px; color:#888; text-align:center;">
    Система учета регистрации заявок
  </p>
</body></html>"""

    try:
        from weasyprint import HTML
        pdf = HTML(string=html).write_pdf()
        media = "application/pdf"
    except ImportError:
        pdf = html.encode("utf-8")
        media = "text/html; charset=utf-8"

    filename = f"services_report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{'pdf' if media == 'application/pdf' else 'html'}"
    return Response(
        content=pdf,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
