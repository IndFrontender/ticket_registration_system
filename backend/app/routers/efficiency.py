import io
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import func

from ..database import get_db
from ..models import Ticket, TicketHistory, UserModel
from .auth import require_admin

router = APIRouter(prefix="/api/efficiency", tags=["efficiency"])


@router.get("/masters")
def get_master_efficiency(
    period: str = Query("month", regex="^(day|month|year)$"),
    master_name: str | None = Query(None),
    db: Session = Depends(get_db),
    admin: UserModel = Depends(require_admin),
):
    today = date.today()
    if period == "day":
        start = today - timedelta(days=30)
        date_trunc = func.date(TicketHistory.created_at)
        label_format = "%d.%m"
    elif period == "month":
        start = today - timedelta(days=365)
        date_trunc = func.strftime("%Y-%m", TicketHistory.created_at)
        label_format = "%m.%Y"
    else:
        start = today - timedelta(days=365 * 3)
        date_trunc = func.strftime("%Y", TicketHistory.created_at)
        label_format = "%Y"

    q = db.query(
        TicketHistory.user,
        date_trunc.label("period"),
        func.count(TicketHistory.id).label("cnt"),
    ).filter(TicketHistory.created_at >= start)

    if master_name:
        q = q.filter(TicketHistory.user == master_name)

    q = q.group_by(TicketHistory.user, date_trunc).order_by(TicketHistory.user, date_trunc)
    rows = q.all()

    masters_set = sorted(set(r.user for r in rows))
    periods_set = sorted(set(r.period for r in rows))

    series = []
    for master in masters_set:
        data = {}
        for r in rows:
            if r.user == master:
                data[str(r.period)] = r.cnt
        series.append({
            "master": master,
            "data": [{"period": p, "count": data.get(p, 0)} for p in periods_set],
            "total": sum(data.values()),
        })

    return {
        "period": period,
        "masters": series,
        "periods": list(periods_set),
        "grand_total": sum(s["total"] for s in series),
    }


@router.get("/masters/export")
def export_master_efficiency(
    period: str = Query("month", regex="^(day|month|year)$"),
    format: str = Query("xlsx"),
    db: Session = Depends(get_db),
    admin: UserModel = Depends(require_admin),
):
    data = get_master_efficiency(period, None, db, admin)

    if format == "xlsx":
        return _export_efficiency_excel(data)
    elif format == "pdf":
        return _export_efficiency_pdf(data)
    return Response("Invalid format", status_code=400)


def _export_efficiency_excel(data: dict) -> Response:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Эффективность"

    header_font = Font(name="DejaVu Sans", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Border(*(Side(style="thin"),) * 4)
    cell_font = Font(name="DejaVu Sans", size=10)

    periods = data["periods"]
    headers = ["Мастер"] + periods + ["Всего"]
    ws.append(headers)
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.font = header_font; c.fill = header_fill; c.alignment = Alignment(horizontal="center"); c.border = thin

    for idx, m in enumerate(data["masters"], 2):
        row = [m["master"]] + [m["data"][i]["count"] if i < len(m["data"]) else 0 for i in range(len(periods))] + [m["total"]]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx, column=col).font = cell_font
            ws.cell(row=idx, column=col).border = thin

    period_label = {"day": "день", "month": "месяц", "year": "год"}.get(data["period"], "")
    filename = f"efficiency_{data['period']}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _export_efficiency_pdf(data: dict) -> Response:
    period_label = {"day": "день", "month": "месяц", "year": "год"}.get(data["period"], "")
    periods = data["periods"]

    rows_html = ""
    for m in data["masters"]:
        cells = f"<td>{m['master']}</td>"
        for p in periods:
            found = [d for d in m["data"] if d["period"] == p]
            cells += f"<td>{found[0]['count'] if found else 0}</td>"
        cells += f"<td><strong>{m['total']}</strong></td>"
        rows_html += f"<tr>{cells}</tr>"

    period_headers = "".join(f"<th>{p}</th>" for p in periods)

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><style>
  @page {{ margin: 15mm; }}
  body {{ font-family: 'DejaVu Sans', Arial, sans-serif; font-size: 12px; }}
  h2 {{ text-align: center; }}
  table {{ width: 100%; border-collapse: collapse; }}
  th, td {{ border: 1px solid #000; padding: 5px 7px; text-align: center; font-size: 10px; }}
  th {{ background: #4472C4; color: #fff; }}
  .total {{ text-align: right; font-size: 12px; margin-top: 8px; }}
  .date {{ text-align: right; font-size: 10px; color: #666; }}
</style></head><body>
  <h2>Эффективность мастеров (по {period_label}ам)</h2>
  <div class="date">Сформирован: {datetime.utcnow().strftime('%d.%m.%Y %H:%M')}</div>
  <table>
    <tr><th>Мастер</th>{period_headers}<th>Всего</th></tr>
    {rows_html}
  </table>
  <div class="total">Общее количество заявок: <strong>{data['grand_total']}</strong></div>
  <p style="margin-top:30px; font-size:10px; color:#888; text-align:center;">
    Система учета регистрации заявок
  </p>
</body></html>"""

    try:
        from weasyprint import HTML
        pdf = HTML(string=html).write_pdf()
        media = "application/pdf"
    except ImportError:
        pdf = html.encode("utf-8")
        media = "text/html; charset=utf-8"

    filename = f"efficiency_{data['period']}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.{'pdf' if media == 'application/pdf' else 'html'}"
    return Response(
        content=pdf,
        media_type=media,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
