import io
from datetime import date, datetime, timedelta
from typing import Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import or_, func

from ..models import Ticket, Client, Equipment, Task, InspectionReport, WarrantyChecklist, TicketService

FIELD_MAPS: dict[str, dict[str, dict]] = {
    "tickets": {
        "number": {"label": "Номер", "getter": lambda t: t.number},
        "client_name": {"label": "Клиент", "getter": lambda t: t.client.name if t.client else ""},
        "client_phone": {"label": "Телефон", "getter": lambda t: t.client.phone if t.client else ""},
        "short_description": {"label": "Описание", "getter": lambda t: t.short_description},
        "priority": {"label": "Приоритет", "getter": lambda t: t.priority.value if hasattr(t.priority, 'value') else str(t.priority)},
        "status": {"label": "Статус", "getter": lambda t: t.status.value if hasattr(t.status, 'value') else str(t.status)},
        "work_date": {"label": "Дата работы", "getter": lambda t: str(t.work_date) if t.work_date else ""},
        "review_date": {"label": "Дата осмотра", "getter": lambda t: str(t.review_date) if t.review_date else ""},
        "warranty_months": {"label": "Гарантия (мес)", "getter": lambda t: t.warranty_months or 0},
        "created_at": {"label": "Создана", "getter": lambda t: t.created_at.strftime('%d.%m.%Y %H:%M') if t.created_at else ""},
        "updated_at": {"label": "Обновлена", "getter": lambda t: t.updated_at.strftime('%d.%m.%Y %H:%M') if t.updated_at else ""},
    },
    "clients": {
        "name": {"label": "Наименование", "getter": lambda c: c.name},
        "client_type": {"label": "Тип", "getter": lambda c: "Юр.лицо" if c.client_type == "legal" else "Физ.лицо"},
        "phone": {"label": "Телефон", "getter": lambda c: c.phone or ""},
        "email": {"label": "Email", "getter": lambda c: c.email or ""},
        "address": {"label": "Адрес", "getter": lambda c: c.address or ""},
        "inn": {"label": "ИНН", "getter": lambda c: c.inn or ""},
        "kpp": {"label": "КПП", "getter": lambda c: c.kpp or ""},
        "contact_person": {"label": "Контактное лицо", "getter": lambda c: c.contact_person or ""},
        "notes": {"label": "Примечания", "getter": lambda c: c.notes or ""},
        "created_at": {"label": "Создан", "getter": lambda c: c.created_at.strftime('%d.%m.%Y') if c.created_at else ""},
        "tickets_count": {"label": "Заявок", "getter": lambda c: len(c.tickets)},
    },
    "equipment": {
        "equipment_type": {"label": "Тип", "getter": lambda e: e.equipment_type},
        "manufacturer": {"label": "Производитель", "getter": lambda e: e.manufacturer or ""},
        "model": {"label": "Модель", "getter": lambda e: e.model or ""},
        "serial_number": {"label": "Серийный номер", "getter": lambda e: e.serial_number or ""},
        "inventory_number": {"label": "Инв. номер", "getter": lambda e: e.inventory_number or ""},
        "location": {"label": "Местоположение", "getter": lambda e: e.location or ""},
        "client_name": {"label": "Клиент", "getter": lambda e: e.client.name if e.client else ""},
        "status": {"label": "Статус", "getter": lambda e: e.status},
        "purchase_date": {"label": "Дата покупки", "getter": lambda e: str(e.purchase_date) if e.purchase_date else ""},
        "warranty_expiry": {"label": "Гарантия до", "getter": lambda e: str(e.warranty_expiry) if e.warranty_expiry else ""},
        "notes": {"label": "Примечания", "getter": lambda e: e.notes or ""},
    },
    "tasks": {
        "title": {"label": "Задача", "getter": lambda t: t.title},
        "description": {"label": "Описание", "getter": lambda t: t.description or ""},
        "category": {"label": "Категория", "getter": lambda t: t.category or ""},
        "priority": {"label": "Приоритет", "getter": lambda t: t.priority},
        "status": {"label": "Статус", "getter": lambda t: t.status},
        "due_date": {"label": "Срок", "getter": lambda t: str(t.due_date) if t.due_date else ""},
        "completed_at": {"label": "Завершена", "getter": lambda t: t.completed_at.strftime('%d.%m.%Y') if t.completed_at else ""},
        "created_at": {"label": "Создана", "getter": lambda t: t.created_at.strftime('%d.%m.%Y') if t.created_at else ""},
    },
    "inspections": {
        "ticket_number": {"label": "Номер заявки", "getter": lambda i: i.ticket.number if i.ticket else ""},
        "report_date": {"label": "Дата проверки", "getter": lambda i: str(i.report_date)},
        "report_type": {"label": "Тип", "getter": lambda i: i.report_type},
        "status": {"label": "Статус", "getter": lambda i: i.status},
        "notes": {"label": "Примечания", "getter": lambda i: i.notes or ""},
        "client_confirmed": {"label": "Подтверждено", "getter": lambda i: "Да" if i.client_confirmed else "Нет"},
        "files_count": {"label": "Файлов", "getter": lambda i: len(i.files)},
    },
    "warranty_checks": {
        "ticket_number": {"label": "Номер заявки", "getter": lambda w: w.ticket.number if w.ticket else ""},
        "check_date": {"label": "Дата проверки", "getter": lambda w: str(w.check_date)},
        "month_number": {"label": "Месяц", "getter": lambda w: w.month_number},
        "status": {"label": "Статус", "getter": lambda w: w.status},
        "notes": {"label": "Примечания", "getter": lambda w: w.notes or ""},
        "files_attached": {"label": "Файлов", "getter": lambda w: w.files_attached or 0},
        "created_at": {"label": "Создана", "getter": lambda w: w.created_at.strftime('%d.%m.%Y') if w.created_at else ""},
        "completed_at": {"label": "Завершена", "getter": lambda w: w.completed_at.strftime('%d.%m.%Y') if w.completed_at else ""},
    },
}

ENTITY_MODELS = {
    "tickets": Ticket,
    "clients": Client,
    "equipment": Equipment,
    "tasks": Task,
    "inspections": InspectionReport,
    "warranty_checks": WarrantyChecklist,
}


def _get_data(entity_type: str, filters: list[dict], db: Session) -> list[Any]:
    model = ENTITY_MODELS.get(entity_type)
    if not model:
        return []
    q = db.query(model)
    for f in filters or []:
        field = f.get("field")
        op = f.get("op", "eq")
        value = f.get("value")
        if not field or not hasattr(model, field):
            continue
        col = getattr(model, field)
        if op == "eq":
            q = q.filter(col == value)
        elif op == "neq":
            q = q.filter(col != value)
        elif op == "gt" and value:
            q = q.filter(col > value)
        elif op == "gte" and value:
            q = q.filter(col >= value)
        elif op == "lt" and value:
            q = q.filter(col < value)
        elif op == "lte" and value:
            q = q.filter(col <= value)
        elif op == "contains" and value:
            q = q.filter(col.ilike(f"%{value}%"))
        elif op == "in" and value:
            q = q.filter(col.in_(value if isinstance(value, list) else [value]))
    return q.order_by(model.id.desc()).all()


def _build_excel(data: list, entity_type: str, columns: list[str], title: str) -> bytes:
    field_map = FIELD_MAPS.get(entity_type, {})
    selected = [c for c in columns if c in field_map]
    if not selected:
        selected = list(field_map.keys())
    labels = [field_map[c]["label"] for c in selected]

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(name="DejaVu Sans", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell_font = Font(name="DejaVu Sans", size=10)
    date_font = Font(name="DejaVu Sans", size=10, color="666666")

    ws.append(labels)
    for col_idx in range(1, len(selected) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, item in enumerate(data, start=2):
        values = []
        for c in selected:
            try:
                v = field_map[c]["getter"](item)
                values.append(v)
            except Exception:
                values.append("")
        ws.append(values)
        for col_idx in range(1, len(selected) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.border = thin_border

    for col_idx, label in enumerate(labels, start=1):
        max_len = len(label)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            for val in row:
                if val:
                    max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def generate_report(entity_type: str, columns: list[str], filters: list[dict], title: str, db: Session) -> bytes:
    data = _get_data(entity_type, filters, db)
    return _build_excel(data, entity_type, columns, title)


def list_available_fields(entity_type: str) -> list[dict]:
    fields = FIELD_MAPS.get(entity_type, {})
    return [{"key": k, "label": v["label"]} for k, v in fields.items()]
